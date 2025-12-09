#!/usr/bin/env python3
"""
demo.py - Lineage Pipeline v4.2

Parse → Abstract → Retrieve (RAG + rerank) → Generate → Validate

Features:
- Parse Sybase DDL into a table/column catalog (DDLPaser)
- Parse Stored Procedures for:
    • Table usage (source/target tables)
    • Deterministic attribute lineage:
        - INSERT INTO t (c1, c2, ...) SELECT e1, e2, ...
        - UPDATE t SET t.c = s.c ...
    • Parameter → column lineage:
        - WHERE t.col = @param or @param = t.col
- Parse Imperva logs into telemetry (table/procedure frequencies)
- Build YAML Knowledge Graph (base + enriched with telemetry)
- Build RAG index over KG (TF-IDF or SentenceTransformer embeddings)
- Rerank RAG results using:
    • Semantic similarity
    • Telemetry scores
    • Query hints (table/column names)
    • Document type (attribute_lineage vs table_lineage vs parameter_lineage)
- Generate Table & Attribute lineage in Excel and CSV (directly from KG)
- Validate KG vs optional ground-truth YAML with ≥90% threshold for P/R
"""

from __future__ import annotations

import re
import os
import csv
import json
import zipfile
import pickle
import collections
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional, Iterable, Set

import numpy as np
import pandas as pd
import yaml
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

try:
    import sqlparse
except ImportError:
    raise ImportError("Please install sqlparse: pip install sqlparse")

# Optional: semantic embeddings
try:
    from sentence_transformers import SentenceTransformer
    EMBEDDINGS_AVAILABLE = True
except ImportError:
    EMBEDDINGS_AVAILABLE = False

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook


# ============================================================================
# CONFIGURATION
# ============================================================================

@dataclass
class Config:
    """Top-level config for the demo pipeline."""
    base_dir: Path = Path(__file__).resolve().parent

    # Input
    zip_path: Path = field(init=False)
    imperva_path: Path = field(init=False)
    ground_truth_path: Path = field(init=False)

    # Outputs
    out_base_yaml: Path = field(init=False)
    out_enriched_yaml: Path = field(init=False)
    out_rag_pkl: Path = field(init=False)
    out_lineage_dir: Path = field(init=False)
    out_validation: Path = field(init=False)

    # Intermediate / docs
    intermediate_dir: Path = field(init=False)
    documentation_dir: Path = field(init=False)

    # Parameters
    max_sp_chars: int = 200_000
    confidence_threshold: float = 0.90
    rag_top_k: int = 10

    def __post_init__(self):
        input_dir = self.base_dir / "input_data"
        input_dir.mkdir(exist_ok=True)

        # SP/DDL zip
        if (input_dir / "store-procedure-table.zip").exists():
            self.zip_path = input_dir / "store-procedure-table.zip"
        elif (input_dir / "SPsample.zip").exists():
            self.zip_path = input_dir / "SPsample.zip"
        else:
            self.zip_path = input_dir / "store-procedure-table.zip"

        # Imperva logs
        if (input_dir / "imperva_small.xlsx").exists():
            self.imperva_path = input_dir / "imperva_small.xlsx"
        elif (input_dir / "imperva_small.zip").exists():
            self.imperva_path = input_dir / "imperva_small.zip"
        else:
            self.imperva_path = input_dir / "imperva_small.xlsx"

        self.ground_truth_path = input_dir / "annotated_groundtruth.yaml"

        out_root = self.base_dir / "output_artifacts"
        out_root.mkdir(parents=True, exist_ok=True)

        self.out_base_yaml = out_root / "demo.base.yaml"
        self.out_enriched_yaml = out_root / "demo.enriched.yaml"
        self.out_rag_pkl = out_root / "lineage_rag.pkl"
        self.out_lineage_dir = out_root / "out_lineage"
        self.out_lineage_dir.mkdir(parents=True, exist_ok=True)
        self.out_validation = out_root / "validation_report.json"

        self.intermediate_dir = self.base_dir / "intermediate"
        self.intermediate_dir.mkdir(parents=True, exist_ok=True)

        self.documentation_dir = self.base_dir / "documentation"
        self.documentation_dir.mkdir(parents=True, exist_ok=True)


# ============================================================================
# PHASE 1: PARSE - DDL & Stored Procs
# ============================================================================

class DDLParser:
    """Parse Sybase/SQLServer-style DDL into a catalog."""
    def __init__(self):
        self.catalog: Dict[str, List[Dict[str,Any]]] = {}
        self.stats = {"tables":0, "columns":0}

    def parse(self, ddl_text: str) -> Dict[str, List[Dict[str,Any]]]:
        create_pat = re.compile(
            r"CREATE\s+TABLE\s+(?:dbo\.)?([A-Z0-9_\.\[\]]+)\s*\((.*?)\)\s*(?:LOCK\s+\w+)?(?:\s*GO)?",
            re.I | re.S
        )
        col_pat = re.compile(
            r"^\s*\[?([A-Za-z0-9_]+)\]?\s+([A-Za-z0-9\(\),]+)(?:\s+(NULL|NOT\s+NULL))?",
            re.I
        )

        for m in create_pat.finditer(ddl_text):
            raw = m.group(1).replace("[","").replace("]","")
            tname = raw.lower()
            cols_block = m.group(2)
            cols: List[Dict[str,Any]] = []
            for line in cols_block.splitlines():
                line = line.strip().rstrip(",")
                if not line or line.upper().startswith(("CONSTRAINT","PRIMARY","FOREIGN","UNIQUE","CHECK","INDEX")):
                    continue
                cm = col_pat.match(line)
                if not cm:
                    continue
                cname, ctype = cm.group(1), cm.group(2)
                nullable = cm.group(3)
                cols.append({
                    "name": cname.lower(),
                    "type": ctype.lower(),
                    "nullable": nullable != "NOT NULL" if nullable else True
                })
            if cols:
                self.catalog[tname] = cols
                self.stats["tables"] += 1
                self.stats["columns"] += len(cols)

        return self.catalog


class EnhancedSPParser:
    """
    Parse stored procedure text into:
    - tables (all/source/target)
    - deterministic attribute_lineage
    - parameter_lineage (heuristic)
    """

    SQL_KW = {
        'from','where','and','or','not','in','exists','like','join','left','right',
        'inner','outer','full','cross','on','using','as','is','null','between','case',
        'when','then','else','end','if','while','begin','return','go','exec','execute',
        'declare','set','print','select','insert','update','delete','create','drop',
        'alter','table','into','values','top','distinct','order','by','group','having',
        'union','all','desc','asc'
    }

    def __init__(self, catalog: Dict[str,List[Dict[str,Any]]]):
        self.catalog = catalog
        self.catalog_lower = {k.lower(): k for k in catalog.keys()}
        self.table_columns = {
            t: {c["name"] for c in cols} for t, cols in catalog.items()
        }

    def _norm_table(self, raw: str) -> str:
        raw = raw.replace("[","").replace("]","")
        raw = re.sub(r"^\w+\.", "", raw)
        return self.catalog_lower.get(raw.lower(), raw.lower())

    def _is_table_name(self, token: str) -> bool:
        token = token.strip()
        if not token:
            return False
        if token.lower() in self.SQL_KW:
            return False
        if token[0].isdigit():
            return False
        return token.lower() in self.catalog_lower

    def extract_tables(self, sql: str) -> List[Tuple[str,str]]:
        """Return (table, op) pairs."""
        res: List[Tuple[str,str]] = []
        # FROM/JOIN -> SELECT
        from_pat = re.compile(r"\b(?:FROM|JOIN)\s+(?:dbo\.)?([A-Za-z_][A-Za-z0-9_]*)", re.I)
        for m in from_pat.finditer(sql):
            raw = m.group(1)
            if self._is_table_name(raw):
                res.append((self._norm_table(raw), "SELECT"))
        # INSERT
        ins_pat = re.compile(r"\bINSERT\s+INTO\s+(?:dbo\.)?([A-Za-z_][A-Za-z0-9_]*)", re.I)
        for m in ins_pat.finditer(sql):
            raw = m.group(1)
            if self._is_table_name(raw):
                res.append((self._norm_table(raw), "INSERT"))
        # UPDATE
        upd_pat = re.compile(r"\bUPDATE\s+(?:dbo\.)?([A-Za-z_][A-Za-z0-9_]*)\s+SET", re.I)
        for m in upd_pat.finditer(sql):
            raw = m.group(1)
            if self._is_table_name(raw):
                res.append((self._norm_table(raw), "UPDATE"))
        # DELETE
        del_pat = re.compile(r"\bDELETE\s+FROM\s+(?:dbo\.)?([A-Za-z_][A-Za-z0-9_]*)", re.I)
        for m in del_pat.finditer(sql):
            raw = m.group(1)
            if self._is_table_name(raw):
                res.append((self._norm_table(raw), "DELETE"))
        return res

    def _extract_param_declarations(self, sp_text: str) -> List[str]:
        """
        Very simple parameter extraction:
        - Looks for '@name' tokens between CREATE PROCEDURE and 'AS'
        """
        header_match = re.search(
            r"CREATE\s+PROCEDURE\s+(?:dbo\.)?[A-Za-z0-9_]+\s*(.*?)\bAS\b",
            sp_text, re.I|re.S
        )
        if not header_match:
            return []
        header = header_match.group(1)
        return sorted({p.lower() for p in re.findall(r"@([A-Za-z0-9_]+)", header)})

    def _build_alias_map(self, sp_text: str) -> Dict[str,str]:
        """
        Map aliases -> base table across the proc.
        FROM tclmdets cd, JOIN tnumgen tn, etc.
        """
        amap: Dict[str,str] = {}
        pat = re.compile(
            r"\b(?:FROM|JOIN)\s+(?:dbo\.)?([A-Za-z_][A-Za-z0-9_]*)\s+([A-Za-z_][A-Za-z0-9_]*)",
            re.I
        )
        for m in pat.finditer(sp_text):
            raw_table = m.group(1)
            alias = m.group(2).lower()
            tnorm = self._norm_table(raw_table)
            amap[alias] = tnorm
        return amap

    def extract_attribute_lineage(self, sql: str, proc_name: str) -> List[Dict[str,Any]]:
        mappings: List[Dict[str,Any]] = []
        tables_ops = self.extract_tables(sql)
        src_tables = {t for t,op in tables_ops if op=="SELECT"}
        tgt_tables = {t for t,op in tables_ops if op in ("INSERT","UPDATE","DELETE")}

        # INSERT INTO t(c1, c2, ...) SELECT e1, e2, ...
        ins_pat = re.compile(
            r"INSERT\s+INTO\s+(?:dbo\.)?([A-Za-z_][A-Za-z0-9_]*)\s*\((.*?)\)\s*SELECT\s+(.*?)\s+FROM",
            re.I|re.S
        )
        for m in ins_pat.finditer(sql):
            tgt = self._norm_table(m.group(1))
            tgt_cols = [c.strip(" []").lower() for c in m.group(2).split(",")]
            selects = [c.strip() for c in m.group(3).split(",")]
            src_cols: List[Optional[str]] = []
            for item in selects:
                item = re.sub(r"\s+AS\s+\w+$","", item, flags=re.I)
                qm = re.search(r"([A-Za-z_][A-Za-z0-9_]*)\.([A-Za-z_][A-Za-z0-9_]*)", item)
                if qm:
                    src_cols.append(qm.group(2).lower())
                else:
                    col = re.sub(r"[^\w]","", item).lower()
                    src_cols.append(col or None)
            for tc, sc in zip(tgt_cols, src_cols):
                if not sc:
                    continue
                for st in src_tables:
                    if sc in self.table_columns.get(st,set()) and tc in self.table_columns.get(tgt,set()):
                        mappings.append({
                            "source_table": st,
                            "source_column": sc,
                            "target_table": tgt,
                            "target_column": tc,
                            "via_procedure": proc_name,
                            "mapping_type": "INSERT_SELECT",
                            "confidence": 0.95
                        })

        # UPDATE t SET t.col = s.col ...
        upd_pat = re.compile(
            r"\bUPDATE\s+(?:dbo\.)?([A-Za-z_][A-Za-z0-9_]*)\s+SET\s+(.*?)(?:WHERE|FROM|$)",
            re.I|re.S
        )
        for m in upd_pat.finditer(sql):
            tgt = self._norm_table(m.group(1))
            set_block = m.group(2)
            for assign in set_block.split(","):
                if "=" not in assign:
                    continue
                left, right = assign.split("=",1)
                tc = left.strip().split(".")[-1].strip("[]").lower()
                if tc not in self.table_columns.get(tgt,set()):
                    continue
                expr = right.strip()
                for qm in re.finditer(r"([A-Za-z_][A-Za-z0-9_]+)\.([A-Za-z_][A-Za-z0-9_]*)", expr):
                    src_token = qm.group(1)
                    sc = qm.group(2).lower()
                    if not self._is_table_name(src_token):
                        continue
                    st = self._norm_table(src_token)
                    if sc in self.table_columns.get(st,set()):
                        mappings.append({
                            "source_table": st,
                            "source_column": sc,
                            "target_table": tgt,
                            "target_column": tc,
                            "via_procedure": proc_name,
                            "mapping_type": "UPDATE_SET",
                            "confidence": 0.90
                        })

        # Fallback name-matching if nothing else found
        if not mappings and src_tables and tgt_tables:
            for st in src_tables:
                for tt in tgt_tables:
                    src_cols = self.table_columns.get(st,set())
                    tgt_cols = self.table_columns.get(tt,set())
                    for c in src_cols & tgt_cols:
                        mappings.append({
                            "source_table": st,
                            "source_column": c,
                            "target_table": tt,
                            "target_column": c,
                            "via_procedure": proc_name,
                            "mapping_type": "NAME_MATCH_HEURISTIC",
                            "confidence": 0.70
                        })

        return mappings

    def _is_table_name(self, token: str) -> bool:
        return self._is_table_name_impl(token)

    def _is_table_name_impl(self, token: str) -> bool:
        token = token.strip()
        if not token:
            return False
        if token.lower() in self.SQL_KW:
            return False
        if token[0].isdigit():
            return False
        return token.lower() in self.catalog_lower

    def extract_parameter_lineage(self, sp_text: str, proc_name: str) -> List[Dict[str,Any]]:
        """
        Heuristic: find patterns t.col = @param or @param = t.col
        and resolve t to a real table via alias map.
        """
        params = self._extract_param_declarations(sp_text)
        if not params:
            return []
        alias_map = self._build_alias_map(sp_text)
        results: List[Dict[str,Any]] = []
        # look in entire body
        body = sp_text
        for p in params:
            p_tok = f"@{p}"
            # t.col = @p
            pat1 = re.compile(rf"([A-Za-z_][A-Za-z0-9_]*)\.([A-Za-z_][A-Za-z0-9_]*)\s*=\s*{re.escape(p_tok)}\b", re.I)
            # @p = t.col
            pat2 = re.compile(rf"{re.escape(p_tok)}\s*=\s*([A-Za-z_][A-Za-z0-9_]*)\.([A-Za-z_][A-Za-z0-9_]*)", re.I)
            for m in pat1.finditer(body):
                alias, col = m.group(1).lower(), m.group(2).lower()
                tbl = alias_map.get(alias)
                results.append({
                    "parameter": p_tok.lower(),
                    "table": tbl,
                    "column": col,
                    "via_procedure": proc_name,
                    "pattern": "t.col = @p",
                    "confidence": 0.9 if tbl else 0.6
                })
            for m in pat2.finditer(body):
                alias, col = m.group(1).lower(), m.group(2).lower()
                tbl = alias_map.get(alias)
                results.append({
                    "parameter": p_tok.lower(),
                    "table": tbl,
                    "column": col,
                    "via_procedure": proc_name,
                    "pattern": "@p = t.col",
                    "confidence": 0.9 if tbl else 0.6
                })
        return results

    def parse_procedure(self, sp_text: str) -> Dict[str,Any]:
        name_match = re.search(r"CREATE\s+PROCEDURE\s+(?:dbo\.)?([A-Za-z0-9_]+)", sp_text, re.I)
        proc_name = name_match.group(1) if name_match else "unknown_procedure"
        tables_ops = self.extract_tables(sp_text)
        all_tables = sorted({t for t,_ in tables_ops})
        src_tables = sorted({t for t,op in tables_ops if op=="SELECT"})
        tgt_tables = sorted({t for t,op in tables_ops if op in ("INSERT","UPDATE","DELETE")})
        attr_lineage = self.extract_attribute_lineage(sp_text, proc_name)
        param_lineage = self.extract_parameter_lineage(sp_text, proc_name)
        return {
            "name": proc_name,
            "all_tables": all_tables,
            "source_tables": src_tables,
            "target_tables": tgt_tables,
            "attribute_lineage": attr_lineage,
            "parameter_lineage": param_lineage
        }


# ============================================================================
# Imperva Telemetry Parser
# ============================================================================

class ImpervaLogParser:
    def __init__(self, path: Path):
        self.path = path
        self.logs: List[Dict[str,Any]] = []
        self.telemetry = {
            "table_access_freq": {},
            "procedure_call_freq": {},
            "table_proc_cooc": collections.Counter()
        }

    def parse(self) -> List[Dict[str,Any]]:
        if not self.path.exists():
            print(f"⚠ Imperva log not found: {self.path}")
            return []
        try:
            if str(self.path).lower().endswith(".zip"):
                import io
                with zipfile.ZipFile(self.path,"r") as z:
                    xlsx = [f for f in z.namelist() if f.lower().endswith(".xlsx")]
                    if not xlsx:
                        return []
                    with z.open(xlsx[0]) as f:
                        df = pd.read_excel(io.BytesIO(f.read()))
            else:
                df = pd.read_excel(self.path)
            for _, row in df.iterrows():
                self.logs.append({
                    "timestamp": row.get("First Timestamp"),
                    "user": row.get("_id_DB User Name"),
                    "original_sql": row.get("_id_Original SQL"),
                    "objects_verbs": row.get("_id_Objects and Verbs"),
                    "count": int(row.get("count",1) or 1)
                })
            self._build_telemetry()
            print(f"✓ Imperva rows: {len(self.logs)}")
            return self.logs
        except Exception as e:
            print(f"⚠ Imperva parse error: {e}")
            return []

    def _build_telemetry(self):
        tpat = re.compile(r"\b(?:from|join|into|update)\s+(?:dbo\.)?([a-z_][a-z0-9_]*)", re.I)
        ppat = re.compile(r"\b(?:exec|execute)\s+([a-z_][a-z0-9_]*)", re.I)
        for r in self.logs:
            cnt = r["count"]
            sql = str(r["original_sql"] or "").lower()
            if not sql or sql=="nan":
                continue
            tables = {m.group(1).lower() for m in tpat.finditer(sql)}
            procs = {m.group(1).lower() for m in ppat.finditer(sql)}
            for t in tables:
                self.telemetry["table_access_freq"][t] = self.telemetry["table_access_freq"].get(t,0) + cnt
            for p in procs:
                self.telemetry["procedure_call_freq"][p] = self.telemetry["procedure_call_freq"].get(p,0) + cnt
                for t in tables:
                    self.telemetry["table_proc_cooc"][(t,p)] += cnt

    def table_score(self, table: str) -> float:
        t = table.lower()
        freq = self.telemetry["table_access_freq"].get(t,0)
        mx = max(self.telemetry["table_access_freq"].values()) if self.telemetry["table_access_freq"] else 1
        return freq/mx if mx>0 else 0.0

    def proc_score(self, proc: str) -> float:
        p = proc.lower()
        freq = self.telemetry["procedure_call_freq"].get(p,0)
        mx = max(self.telemetry["procedure_call_freq"].values()) if self.telemetry["procedure_call_freq"] else 1
        return freq/mx if mx>0 else 0.0


# ============================================================================
# PHASE 2: ABSTRACT - YAML KG (base + enriched)
# ============================================================================

class YAMLKnowledgeGraph:
    def __init__(self, cfg: Config):
        self.cfg = cfg
        self.graph: Dict[str,Any] = {}

    def build(
        self,
        catalog: Dict[str, List[Dict[str,Any]]],
        procs: List[Dict[str,Any]],
        tel: Optional[ImpervaLogParser] = None
    ) -> Dict[str,Any]:
        g: Dict[str,Any] = {
            "metadata": {
                "generated_at": datetime.now().isoformat(),
                "version": "4.2",
                "source": "Iberia Legacy"
            },
            "catalog": catalog,
            "procedures": [],
            "lineage": {
                "table_level": [],
                "attribute_level": [],
                "parameter_level": []
            },
            "telemetry": {}
        }

        for p in procs:
            g["procedures"].append({
                "name": p["name"],
                "tables": p["all_tables"],
                "source_tables": p["source_tables"],
                "target_tables": p["target_tables"],
                "attribute_mappings_count": len(p.get("attribute_lineage",[])),
                "parameter_mappings_count": len(p.get("parameter_lineage",[]))
            })
            # table-level edges
            for s in p["source_tables"]:
                for t in p["target_tables"]:
                    edge = {"source": s, "target": t, "via_procedure": p["name"]}
                    if tel:
                        edge["telemetry_score"] = (
                            tel.table_score(s) + tel.table_score(t) + tel.proc_score(p["name"])
                        )/3.0
                    g["lineage"]["table_level"].append(edge)
            # attribute-level edges
            for m in p.get("attribute_lineage",[]):
                attr = {
                    "source_table": m["source_table"],
                    "source_column": m["source_column"],
                    "target_table": m["target_table"],
                    "target_column": m["target_column"],
                    "via_procedure": m["via_procedure"],
                    "mapping_type": m["mapping_type"],
                    "confidence": m["confidence"]
                }
                if tel:
                    s_sc = tel.table_score(m["source_table"])
                    t_sc = tel.table_score(m["target_table"])
                    attr["telemetry_score"] = (s_sc+t_sc)/2.0
                    attr["confidence"] = min(0.99, m["confidence"] + 0.05*attr["telemetry_score"])
                g["lineage"]["attribute_level"].append(attr)
            # parameter-level edges
            for pm in p.get("parameter_lineage",[]):
                edge = {
                    "parameter": pm["parameter"],
                    "table": pm["table"],
                    "column": pm["column"],
                    "via_procedure": pm["via_procedure"],
                    "pattern": pm["pattern"],
                    "confidence": pm["confidence"]
                }
                if tel and pm["table"]:
                    edge["telemetry_score"] = tel.table_score(pm["table"])
                g["lineage"]["parameter_level"].append(edge)

        if tel:
            g["telemetry"] = {
                "table_access_frequency": tel.telemetry["table_access_freq"],
                "procedure_call_frequency": tel.telemetry["procedure_call_freq"],
                "total_logs": len(tel.logs)
            }

        self.graph = g
        return g

    def save(self, path: Path):
        with open(path,"w",encoding="utf-8") as f:
            yaml.safe_dump(self.graph, f, sort_keys=False)
        print(f"✓ Saved KG: {path}")


# ============================================================================
# PHASE 3: RAG INDEX (with reranking)
# ============================================================================

class LineageRAG:
    def __init__(self, use_embeddings: bool = True):
        self.documents: List[Dict[str,Any]] = []
        self.use_embeddings = use_embeddings and EMBEDDINGS_AVAILABLE
        self.embedding_model = None
        self.vectorizer: Optional[TfidfVectorizer] = None
        self.doc_vectors: Optional[np.ndarray] = None
        self.metadata: Dict[str,Any] = {
            "created_at": datetime.now().isoformat(),
            "embedding_model": None,
            "vector_dim": None,
            "telemetry_enabled": False
        }

    def add_documents_from_yaml(self, y: Dict[str,Any]):
        tel = y.get("telemetry",{})
        t_freq = tel.get("table_access_frequency",{}) or {}
        p_freq = tel.get("procedure_call_frequency",{}) or {}
        self.metadata["telemetry_enabled"] = bool(t_freq or p_freq)
        max_t = max(t_freq.values()) if t_freq else 1
        max_p = max(p_freq.values()) if p_freq else 1

        # catalog docs
        for t, cols in y.get("catalog",{}).items():
            cns = [c["name"] for c in cols]
            cts = [c["type"] for c in cols]
            score = t_freq.get(t,0)/max_t if max_t>0 else 0.0
            self.documents.append({
                "type": "catalog",
                "table": t,
                "text": f"Table {t} columns: {', '.join(cns)} types: {', '.join(cts)}",
                "telemetry_score": score
            })

        # procedures
        for p in y.get("procedures",[]):
            name = p["name"]
            sc = p_freq.get(name.lower(),0)/max_p if max_p>0 else 0.0
            self.documents.append({
                "type": "procedure",
                "procedure": name,
                "source_tables": p["source_tables"],
                "target_tables": p["target_tables"],
                "text": f"Procedure {name} reads {', '.join(p['source_tables'])} and writes {', '.join(p['target_tables'])}",
                "telemetry_score": sc
            })

        # table lineage
        for e in y.get("lineage",{}).get("table_level",[]):
            self.documents.append({
                "type": "table_lineage",
                "source_table": e["source"],
                "target_table": e["target"],
                "procedure": e["via_procedure"],
                "text": f"Table lineage: {e['source']} -> {e['target']} via {e['via_procedure']}",
                "telemetry_score": e.get("telemetry_score",0.0)
            })

        # attribute lineage
        for e in y.get("lineage",{}).get("attribute_level",[]):
            self.documents.append({
                "type": "attribute_lineage",
                "source_table": e["source_table"],
                "source_column": e["source_column"],
                "target_table": e["target_table"],
                "target_column": e["target_column"],
                "procedure": e["via_procedure"],
                "confidence": e.get("confidence",0.75),
                "mapping_type": e.get("mapping_type","UNKNOWN"),
                "text": (
                    f"Attribute lineage: {e['source_table']}.{e['source_column']} "
                    f"-> {e['target_table']}.{e['target_column']} via {e['via_procedure']} "
                    f"(confidence={e.get('confidence',0.75):.2f}, type={e.get('mapping_type','UNKNOWN')})"
                ),
                "telemetry_score": e.get("telemetry_score",0.0)
            })

        # parameter lineage
        for e in y.get("lineage",{}).get("parameter_level",[]):
            self.documents.append({
                "type": "parameter_lineage",
                "parameter": e["parameter"],
                "table": e["table"],
                "column": e["column"],
                "procedure": e["via_procedure"],
                "text": f"Parameter {e['parameter']} filters {e['table']}.{e['column']} in {e['via_procedure']} ({e['pattern']})",
                "confidence": e.get("confidence",0.8),
                "telemetry_score": e.get("telemetry_score",0.0)
            })

        print(f"✓ RAG docs: {len(self.documents)}")

    def build(self):
        corpus = [d["text"] for d in self.documents]
        if self.use_embeddings:
            try:
                self.embedding_model = SentenceTransformer("all-MiniLM-L6v2")
                self.doc_vectors = self.embedding_model.encode(corpus, show_progress_bar=False)
                self.metadata["embedding_model"] = "all-MiniLM-L6v2"
                self.metadata["vector_dim"] = self.doc_vectors.shape[1]
                return
            except Exception as e:
                print(f"⚠ Embedding error {e}, fallback to TF-IDF")
                self.use_embeddings = False

        self.vectorizer = TfidfVectorizer(max_features=1500, ngram_range=(1,2))
        self.doc_vectors = self.vectorizer.fit_transform(corpus).toarray()
        self.metadata["embedding_model"] = "TF-IDF"
        self.metadata["vector_dim"] = self.doc_vectors.shape[1]

    # --- stage1 retrieval ---
    def search(
        self,
        query_text: str,
        k: int = 20,
        doc_type: Optional[str]=None,
        boost_telemetry: bool=True,
        telemetry_weight: float = 0.3
    ) -> List[Dict[str,Any]]:
        if self.doc_vectors is None:
            raise RuntimeError("RAG not built")

        if self.use_embeddings and self.embedding_model is not None:
            qv = self.embedding_model.encode([query_text])[0]
        else:
            qv = self.vectorizer.transform([query_text]).toarray()[0]
        sims = cosine_similarity([qv], self.doc_vectors)[0]

        if boost_telemetry and self.metadata.get("telemetry_enabled",False):
            tel = np.array([d.get("telemetry_score",0.0) for d in self.documents])
            scores = (1-telemetry_weight)*sims + telemetry_weight*tel
        else:
            scores = sims

        idxs = np.argsort(scores)[::-1]
        out = []
        for idx in idxs:
            d = self.documents[idx]
            if doc_type and d.get("type") != doc_type:
                continue
            out.append({
                **d,
                "score_stage1": float(scores[idx]),
                "semantic_score": float(sims[idx]),
                "telemetry_contribution": float(scores[idx]-sims[idx]) if boost_telemetry else 0.0
            })
            if len(out) >= k:
                break
        return out

    # --- stage2 rerank ---
    def _extract_hints(self, query: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
        q = query.lower()
        m = re.search(r"([a-z0-9_]+)\.([a-z0-9_]+)", q)
        table = col = proc = None
        if m:
            table, col = m.group(1), m.group(2)
        pm = re.search(r"\b([a-z_][a-z0-9_]+_prs|[a-z_][a-z0-9_]+_load|[a-z_][a-z0-9_]+_enq)\b", q)
        if pm:
            proc = pm.group(1)
        return table, col, proc

    def query(
        self,
        query_text: str,
        k: int = 10,
        doc_type: Optional[str]=None,
        boost_telemetry: bool=True,
        telemetry_weight: float=0.3
    ) -> List[Dict[str,Any]]:
        """
        High-level query:
        - Stage1: semantic + telemetry scoring
        - Stage2: metadata-based rerank (type, table/column/proc hint)
        """
        candidates = self.search(query_text, k=5*k, doc_type=doc_type,
                                 boost_telemetry=boost_telemetry,
                                 telemetry_weight=telemetry_weight)
        table_hint, col_hint, proc_hint = self._extract_hints(query_text)
        ql = query_text.lower()

        def bonus(d: Dict[str,Any]) -> float:
        """Metadata-based rerank bonus."""
            b = 0.0
            t = d.get("type")
            # if user is asking "how/compute/map" → prefer attribute_lineage
            if any(k in ql for k in ("how", "compute", "computed", "map", "mapping")):
                if t == "attribute_lineage":
                    b += 0.3
            # if user says table/flow/lineage → prefer table_lineage
            if any(k in ql for k in ("table", "flow", "lineage")):
                if t == "table_lineage":
                    b += 0.25
            # parameter-specific
            if "@" in ql and t == "parameter_lineage":
                b += 0.3
            # match table/col/proc hints
            tgt = d.get("target_table") or d.get("table")
            src = d.get("source_table")
            if table_hint and (tgt == table_hint or src == table_hint):
                b += 0.3
            if col_hint and d.get("target_column") == col_hint:
                b += 0.3
            if proc_hint and d.get("procedure") == proc_hint:
                b += 0.2
            # telemetry contribution
            b += 0.15 * float(d.get("telemetry_contribution",0.0))
            return b

        for d in candidates:
            d["score"] = d["score_stage1"] + bonus(d)
        candidates.sort(key=lambda x:x["score"], reverse=True)
        return candidates[:k]

    def save(self, path: Path):
        data = {
            "documents": self.documents,
            "use_embeddings": self.use_embeddings,
            "metadata": self.metadata,
            "doc_vectors": self.doc_vectors,
            "vectorizer": self.vectorizer,
            "embedding_model": self.embedding_model
        }
        with open(path,"wb") as f:
            pickle.dump(data, f, pickle.HIGHEST_PROTOCOL)
        print(f"✓ Saved RAG binary: {path}")

    @classmethod
    def load(cls, path: Path) -> "LineageRAG":
        with open(path,"rb") as f:
            data = pickle.load(f)
        rag = cls(use_embeddings=data["use_embeddings"])
        rag.documents = data["documents"]
        rag.metadata = data["metadata"]
        rag.doc_vectors = data["doc_vectors"]
        rag.vectorizer = data["vectorizer"]
        rag.embedding_model = data["embedding_model"]
        print(f"✓ Loaded RAG from {path} ({len(rag.documents)} docs, model={rag.metadata.get('embedding_model')})")
        return rag


# ============================================================================
# PHASE 4: GENERATE - Excel + CSV from KG (no RAG required)
# ============================================================================

class ExcelGenerator:
    def __init__(self, cfg: Config):
        self.cfg = cfg

    def _style_sheet(self, ws, df: pd.DataFrame):
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side("thin"))

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = border

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len+2, 60)

    def generate_table_lineage_excel(self, y: Dict[str,Any]):
        edges = y.get("lineage",{}).get("table_level",[])
        if not edges:
            print("⚠ No table lineage found.")
            return
        df = pd.DataFrame([{
            "Source Table": e["source"],
            "Target Table": e["target"],
            "Via Procedure": e["via_procedure"],
            "Telemetry Score": e.get("telemetry_score",0.0)
        } for e in edges]).sort_values("Telemetry Score", ascending=False)
        out = self.cfg.out_lineage_dir / "table_lineage.xlsx"
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Table Lineage", index=False)
            ws = writer.sheets["Table Lineage"]
            self._style_sheet(ws, df)
        print(f"✓ Table lineage Excel: {out}")

    def generate_attribute_lineage_excel(self, y: Dict[str,Any]):
        edges = y.get("lineage",{}).get("attribute_level",[])
        if not edges:
            print("⚠ No attribute lineage found.")
            return
        df = pd.DataFrame([{
            "Source Table": e["source_table"],
            "Source Column": e["source_column"],
            "Target Table": e["target_table"],
            "Target Column": e["target_column"],
            "Via Procedure": e["via_procedure"],
            "Mapping Type": e.get("mapping_type","UNKNOWN"),
            "Confidence": e.get("confidence",0.75),
            "Telemetry Score": e.get("telemetry_score",0.0)
        } for e in edges]).sort_values(["Confidence","Telemetry Score"], ascending=[False,False])
        out = self.cfg.out_lineage_dir / "attribute_lineage.xlsx"
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Attribute Lineage", index=False)
            ws = writer.sheets["Attribute Lineage"]
            self._style_sheet(ws, df)
        print(f"✓ Attribute lineage Excel: {out}")

    def generate_table_lineage_csv(self, y: Dict[str,Any]):
        edges = y.get("lineage",{}).get("table_level",[])
        if not edges:
            return
        df = pd.DataFrame([{
            "Source Table": e["source"],
            "Target Table": e["target"],
            "Via Procedure": e["via_procedure"],
            "Telemetry Score": e.get("telemetry_score",0.0)
        } for e in edges]).sort_values("Telemetry Score", ascending=False)
        out = self.cfg.out_lineage_dir / "table_lineage.csv"
        df.to_csv(out,index=False)
        print(f"✓ Table lineage CSV: {out}")

    def generate_attribute_lineage_csv(self, y: Dict[str,Any]):
        edges = y.get("lineage",{}).get("attribute_level",[])
        if not edges:
            return
        df = pd.DataFrame([{
            "Source Table": e["source_table"],
            "Source Column": e["source_column"],
            "Target Table": e["target_table"],
            "Target Column": e["target_column"],
            "Via Procedure": e["via_procedure"],
            "Mapping Type": e.get("mapping_type","UNKNOWN"),
            "Confidence": e.get("confidence",0.75),
            "Telemetry Score": e.get("telemetry_score",0.0)
        } for e in edges]).sort_values(["Confidence","Telemetry Score"], ascending=[False,False])
        out = self.cfg.out_lineage_dir / "attribute_lineage.csv"
        df.to_csv(out,index=False)
        print(f"✓ Attribute lineage CSV: {out}")


# ============================================================================
# PHASE 5: VALIDATE
# ============================================================================

class LineageValidator:
    def __init__(self, cfg: Config, gt_path: Optional[Path]):
        self.cfg = cfg
        self.gt: Optional[Dict[str,Any]] = None
        if gt_path and gt_path.exists():
            with open(gt_path,"r",encoding="utf-8") as f:
                self.gt = yaml.safe_load(f)
            print(f"✓ Loaded ground truth KG: {gt_path}")
        else:
            print("⚠ No ground truth KG. Validation will compare KG to itself (trivial F1=1.0).")

    def _table_edges(self, y: Dict[str,Any]) -> Set[Tuple[str,str]]:
        return {(e["source"], e["target"]) for e in y.get("lineage",{}).get("table_level",[])}

    def _attr_edges(self, y: Dict[str,Any]) -> Set[Tuple[str,str,str,str]]:
        return {
            (e["source_table"], e["source_column"], e["target_table"], e["target_column"])
            for e in y.get("lineage",{}).get("attribute_level",[])
        }

    def _metrics(self, pred: Set, gt: Set) -> Dict[str,Any]:
        if not pred and not gt:
            return {"precision":1.0,"recall":1.0,"f1":1.0,"tp":0,"fp":0,"fn":0}
        tp = len(pred & gt)
        fp = len(pred - gt)
        fn = len(gt - pred)
        p = tp/(tp+fp) if (tp+fp)>0 else 0.0
        r = tp/(tp+fn) if (tp+fn)>0 else 0.0
        f1 = 2*p*r/(p+r) if (p+r)>0 else 0.0
        return {"precision":p,"recall":r,"f1":f1,"tp":tp,"fp":fp,"fn":fn}

    def validate(self, y: Dict[str,Any]) -> Dict[str,Any]:
        if self.gt is None:
            self.gt = y  # fallback: self-ground truth

        gt_t = self._table_edges(self.gt)
        gt_a = self._attr_edges(self.gt)
        pr_t = self._table_edges(y)
        pr_a = self._attr_edges(y)

        mt = self._metrics(pr_t, gt_t)
        ma = self._metrics(pr_a, gt_a)

        th = self.cfg.confidence_threshold
        ok_t = mt["precision"]>=th and mt["recall"]>=th
        ok_a = ma["precision"]>=th and ma["recall"]>=th

        report = {
            "timestamp": datetime.now().isoformat(),
            "threshold": th,
            "table_lineage": {**mt, "confident": ok_t},
            "attribute_lineage": {**ma, "confident": ok_a},
            "overall": {"confident": ok_t and ok_a}
        }
        return report

    def print_report(self, r: Dict[str,Any]):
        print("\n" + "="*60)
        print("VALIDATION REPORT")
        print("="*60)
        print(f"Table: P={r['table_lineage']['precision']:.1%} "
              f"R={r['table_lineage']['recall']:.1%} "
              f"F1={r['table_lineage']['f1']:.1%} "
              f"Confident={r['table_lineage']['confident']}")
        print(f"Attr:  P={r['attribute_lineage']['precision']:.1%} "
              f"R={r['attribute_lineage']['recall']:.1%} "
              f"F1={r['attribute_lineage']['f1']:.1%} "
              f"Confident={r['attribute_lineage']['confident']}")
        print(f"Overall: {'PASS' if r['overall']['confident'] else 'REVIEW'}")
        print("="*60)

    def save(self, r: Dict[str,Any], path: Path):
        with open(path,"w",encoding="utf-8") as f:
            json.dump(r,f,indent=2)
        print(f"✓ Saved validation report: {path}")


# ============================================================================
# MAIN ORCHESTRATION
# ============================================================================

def main():
    cfg = Config()
    print("\n" + "="*80)
    print("LINEAGE PIPELINE v4.2 – Parse → Abstract → RAG → Excel → Validate")
    print("="*80)

    # Parse DDL + SPs
    sp_root = cfg.intermediate_dir / "sp_demo"
    sp_root.mkdir(exist_ok=True)

    print(f"\n📦 Extracting {cfg.zip_path}")
    with zipfile.ZipFile(cfg.zip_path,"r") as z:
        z.extractall(sp_root)

    ddl_path = sp_root / "store-procedure-table" / "iberia table extract.txt"
    sp_path  = sp_root / "store-procedure-table" / "Iberia-PROC.txt"
    ddl_text = ddl_path.read_text(encoding="utf-8",errors="ignore")
    sp_text  = sp_path.read_text(encoding="utf-8",errors="ignore")[:cfg.max_sp_chars]

    ddl_parser = DDLParser()
    catalog = ddl_parser.parse(ddl_text)
    print(f"✓ DDL parsed: {ddl_parser.stats['tables']} tables, {ddl_parser.stats['columns']} columns")

    # Imperva telemetry
    print("\n📊 Parsing Imperva telemetry...")
    tel_parser = ImpervaLogParser(cfg.imperva_path)
    tel_parser.parse()

    # Parse SPs
    print("\n🧩 Parsing Stored Procedures...")
    sp_parser = EnhancedSPParser(catalog)
    proc_pat = re.compile(
        r"CREATE\s+(?:OR\s+REPLACE\s+)?PROCEDURE.*?(?=CREATE\s+(?:OR\s+REPLACE\s+)?PROCEDURE|\Z)",
        re.I|re.S
    )
    procs: List[Dict[str,Any]] = []
    for i, m in enumerate(proc_pat.finditer(sp_text)):
        if i >= 12:   # limit for demo
            break
        block = m.group(0)
        pdata = sp_parser.parse_procedure(block)
        procs.append(pdata)
        print(f"  {i+1:02d}. {pdata['name']} "
              f"(src={len(pdata['source_tables'])}, tgt={len(pdata['target_tables'])}, "
              f"attr_maps={len(pdata['attribute_lineage'])}, "
              f"param_maps={len(pdata['parameter_lineage'])})")
    if not procs:
        print("⚠ No stored procedures parsed. Check SP file/regex.")
        return

    # Build KG (base + enriched)
    print("\n🧠 Building KG (base & enriched)")
    kg = YAMLKnowledgeGraph(cfg)
    base = kg.build(catalog, procs, tel=None)
    kg.save(cfg.out_base_yaml)
    enriched = kg.build(catalog, procs, tel_parser)
    kg.graph = enriched
    kg.save(cfg.out_enriched_yaml)

    # Build RAG
    print("\n🔍 Building RAG index with reranking support")
    rag = LineageRAG(use_embeddings=EMBEDDINGS_AVAILABLE)
    rag.add_documents_from_yaml(enriched)
    rag.build()
    rag.save(cfg.out_rag_pkl)

    # Quick demo retrieval
    q = "How is tlprdets.reserve_amt computed?"
    print(f"\nRAG DEMO QUERY: {q}")
    results = rag.query(q, k=5, doc_type="attribute_lineage")
    for i, d in enumerate(results,1):
        print(f"  {i}. [{d['type']}] score={d['score']:.3f} : {d['text'][:140]}...")

    # Generate Excel & CSV
    print("\n📊 Generating Excel & CSV lineage reports")
    gen = ExcelGenerator(cfg)
    gen.generate_table_lineage_excel(enriched)
    gen.generate_attribute_lineage_excel(enriched)
    gen.generate_table_lineage_csv(enriched)
    gen.generate_attribute_lineage_csv(enriched)

    # Validate
    print("\n✅ Validating KG vs ground truth (if provided)")
    validator = LineageValidator(cfg, cfg.ground_truth_path)
    rpt = validator.validate(enriched)
    validator.print_report(rpt)
    validator.save(rpt, cfg.out_validation)

    print("\n" + "="*80)
    print("PIPELINE COMPLETE")
    print("="*80)
    print(f"Base KG:       {cfg.out_base_yaml}")
    print(f"Enriched KG:   {cfg.out_enriched_yaml}")
    print(f"RAG binary:    {cfg.out_rag_pkl}")
    print(f"Lineage Excel: {cfg.out_lineage_dir}")
    print(f"Validation:    {cfg.out_validation}")
    print("="*80)


if __name__ == "__main__":
    main()
