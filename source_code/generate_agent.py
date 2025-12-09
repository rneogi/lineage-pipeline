"""
generate_agent.py - PHASE 4: GENERATE Agent (v4.2)

Generates CSV and Excel reports for table and attribute lineage.

Features (v4.2):
- Table lineage reports (sorted by telemetry score)
- Attribute lineage reports (sorted by confidence + telemetry)
- Conditional formatting for confidence scores
- Mapping type breakdown statistics

Usage:
    python generate_agent.py

Input:
    intermediate/retrieve_output.json - Output from Retrieve Agent

Outputs:
    output_artifacts/out_lineage/table_lineage.csv
    output_artifacts/out_lineage/attribute_lineage.csv
    output_artifacts/out_lineage/table_lineage.xlsx
    output_artifacts/out_lineage/attribute_lineage.xlsx
    intermediate/generate_output.json - JSON summary for next stage
"""

import json
from pathlib import Path
from typing import Dict, Any, List
from datetime import datetime
from dataclasses import dataclass, field

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule


# ============================================================================
# CONFIGURATION
# ============================================================================

@dataclass
class Config:
    """Pipeline configuration"""
    base_dir: Path = field(default_factory=lambda: Path(__file__).resolve().parent.parent)

    # Input paths
    retrieve_output: Path = field(init=False)

    # Output paths
    intermediate_dir: Path = field(init=False)
    output_artifacts_dir: Path = field(init=False)
    out_lineage_dir: Path = field(init=False)
    generate_output: Path = field(init=False)

    def __post_init__(self):
        self.intermediate_dir = self.base_dir / "intermediate"
        self.output_artifacts_dir = self.base_dir / "output_artifacts"
        self.out_lineage_dir = self.output_artifacts_dir / "out_lineage"

        self.retrieve_output = self.intermediate_dir / "retrieve_output.json"
        self.generate_output = self.intermediate_dir / "generate_output.json"

        # Create directories
        self.intermediate_dir.mkdir(parents=True, exist_ok=True)
        self.output_artifacts_dir.mkdir(parents=True, exist_ok=True)
        self.out_lineage_dir.mkdir(parents=True, exist_ok=True)


# ============================================================================
# REPORT GENERATOR
# ============================================================================

class ReportGenerator:
    """Generate CSV and Excel reports for lineage data"""

    def __init__(self, config: Config):
        self.config = config

    def generate_table_lineage_csv(self, yaml_data: Dict[str, Any]) -> str:
        """Generate table lineage CSV"""
        edges = yaml_data.get('lineage', {}).get('table_level', [])

        if not edges:
            print("Warning: No table lineage edges to export to CSV")
            return None

        df = pd.DataFrame([
            {
                'Source Table': e['source'],
                'Target Table': e['target'],
                'Via Procedure': e['via_procedure'],
                'Telemetry Score': e.get('telemetry_score', 0.0)
            }
            for e in edges
        ])

        df = df.sort_values('Telemetry Score', ascending=False)

        output_path = self.config.out_lineage_dir / "table_lineage.csv"
        df.to_csv(output_path, index=False)
        print(f"Generated table lineage CSV: {output_path} ({len(df)} edges)")
        return str(output_path)

    def generate_attribute_lineage_csv(self, yaml_data: Dict[str, Any]) -> str:
        """Generate attribute lineage CSV"""
        edges = yaml_data.get('lineage', {}).get('attribute_level', [])

        if not edges:
            print("Warning: No attribute lineage edges to export to CSV")
            return None

        df = pd.DataFrame([
            {
                'Source Table': e['source_table'],
                'Source Column': e['source_column'],
                'Target Table': e['target_table'],
                'Target Column': e['target_column'],
                'Via Procedure': e['via_procedure'],
                'Confidence': e.get('confidence', 0.75),
                'Mapping Type': e.get('mapping_type', 'UNKNOWN'),
                'Telemetry Score': e.get('telemetry_score', 0.0)
            }
            for e in edges
        ])

        df = df.sort_values(['Confidence', 'Telemetry Score'], ascending=[False, False])

        output_path = self.config.out_lineage_dir / "attribute_lineage.csv"
        df.to_csv(output_path, index=False)
        print(f"Generated attribute lineage CSV: {output_path} ({len(df)} edges)")
        return str(output_path)

    def generate_table_lineage_excel(self, yaml_data: Dict[str, Any]) -> str:
        """Generate table lineage Excel with formatting"""
        edges = yaml_data.get('lineage', {}).get('table_level', [])

        if not edges:
            print("Warning: No table lineage edges to export to Excel")
            return None

        df = pd.DataFrame([
            {
                'Source Table': e['source'],
                'Target Table': e['target'],
                'Via Procedure': e['via_procedure'],
                'Telemetry Score': e.get('telemetry_score', 0.0)
            }
            for e in edges
        ])

        df = df.sort_values('Telemetry Score', ascending=False)

        output_path = self.config.out_lineage_dir / "table_lineage.xlsx"
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Table Lineage', index=False)
            ws = writer.sheets['Table Lineage']
            self._apply_formatting(ws, df)

        print(f"Generated table lineage Excel: {output_path} ({len(df)} edges)")
        return str(output_path)

    def generate_attribute_lineage_excel(self, yaml_data: Dict[str, Any]) -> str:
        """Generate attribute lineage Excel with formatting and conditional colors"""
        edges = yaml_data.get('lineage', {}).get('attribute_level', [])

        if not edges:
            print("Warning: No attribute lineage edges to export to Excel")
            return None

        df = pd.DataFrame([
            {
                'Source Table': e['source_table'],
                'Source Column': e['source_column'],
                'Target Table': e['target_table'],
                'Target Column': e['target_column'],
                'Via Procedure': e['via_procedure'],
                'Confidence': e.get('confidence', 0.75),
                'Mapping Type': e.get('mapping_type', 'UNKNOWN'),
                'Telemetry Score': e.get('telemetry_score', 0.0)
            }
            for e in edges
        ])

        df = df.sort_values(['Confidence', 'Telemetry Score'], ascending=[False, False])

        output_path = self.config.out_lineage_dir / "attribute_lineage.xlsx"
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Attribute Lineage', index=False)
            ws = writer.sheets['Attribute Lineage']
            self._apply_formatting(ws, df)

            # Conditional formatting for confidence column
            confidence_col = df.columns.get_loc('Confidence') + 1
            col_letter = chr(64 + confidence_col)
            ws.conditional_formatting.add(
                f'{col_letter}2:{col_letter}{len(df) + 1}',
                ColorScaleRule(
                    start_type='percentile', start_value=10, start_color='F8696B',
                    mid_type='percentile', mid_value=50, mid_color='FFEB84',
                    end_type='percentile', end_value=90, end_color='63BE7B'
                )
            )

        print(f"Generated attribute lineage Excel: {output_path} ({len(df)} edges)")

        # Print mapping type statistics
        print("  Mapping types:")
        for mapping_type, count in df['Mapping Type'].value_counts().items():
            print(f"    {mapping_type}: {count}")

        return str(output_path)

    def _apply_formatting(self, ws, df):
        """Apply Excel formatting"""
        # Header styling
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-width columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border

    def get_statistics(self, yaml_data: Dict[str, Any]) -> Dict[str, Any]:
        """Get statistics about generated reports"""
        table_edges = yaml_data.get('lineage', {}).get('table_level', [])
        attr_edges = yaml_data.get('lineage', {}).get('attribute_level', [])

        # Mapping type breakdown
        mapping_types = {}
        for edge in attr_edges:
            mtype = edge.get('mapping_type', 'UNKNOWN')
            mapping_types[mtype] = mapping_types.get(mtype, 0) + 1

        # Confidence statistics
        confidences = [e.get('confidence', 0) for e in attr_edges]
        avg_confidence = sum(confidences) / len(confidences) if confidences else 0

        return {
            'table_edges': len(table_edges),
            'attribute_edges': len(attr_edges),
            'mapping_types': mapping_types,
            'avg_confidence': avg_confidence,
            'high_confidence_count': sum(1 for c in confidences if c >= 0.90)
        }


# ============================================================================
# MAIN GENERATE AGENT
# ============================================================================

def main():
    """Execute Generate Agent"""
    print("=" * 70)
    print("PHASE 4: GENERATE AGENT")
    print("=" * 70)
    print("Generating CSV and Excel Reports")
    print("-" * 70)

    config = Config()

    # Load Retrieve Agent output
    print(f"\nLoading Retrieve output from {config.retrieve_output}...")
    if not config.retrieve_output.exists():
        print(f"ERROR: Retrieve output not found: {config.retrieve_output}")
        print("Please run retrieve_agent.py first.")
        return

    with open(config.retrieve_output, 'r') as f:
        retrieve_data = json.load(f)

    yaml_graph = retrieve_data['yaml_graph']
    print(f"Loaded YAML graph with {len(yaml_graph.get('lineage', {}).get('table_level', []))} table edges, "
          f"{len(yaml_graph.get('lineage', {}).get('attribute_level', []))} attribute edges")

    # Generate reports
    generator = ReportGenerator(config)

    print("\nGenerating reports...")

    # CSV reports
    table_csv = generator.generate_table_lineage_csv(yaml_graph)
    attr_csv = generator.generate_attribute_lineage_csv(yaml_graph)

    # Excel reports
    table_xlsx = generator.generate_table_lineage_excel(yaml_graph)
    attr_xlsx = generator.generate_attribute_lineage_excel(yaml_graph)

    # Get statistics
    stats = generator.get_statistics(yaml_graph)

    # Save generate output for next stage
    generate_output = {
        'phase': 'GENERATE',
        'timestamp': datetime.now().isoformat(),
        'version': '4.2',
        'input_file': str(config.retrieve_output),
        'outputs': {
            'table_lineage_csv': table_csv,
            'attribute_lineage_csv': attr_csv,
            'table_lineage_xlsx': table_xlsx,
            'attribute_lineage_xlsx': attr_xlsx
        },
        'stats': stats,
        'yaml_graph': yaml_graph  # Pass through for Validate agent
    }

    with open(config.generate_output, 'w') as f:
        json.dump(generate_output, f, indent=2, default=str)

    print("\n" + "=" * 70)
    print("GENERATE AGENT COMPLETE (v4.2)")
    print("=" * 70)
    print("\nStatistics:")
    print(f"  Table Edges:          {stats['table_edges']}")
    print(f"  Attribute Edges:      {stats['attribute_edges']}")
    print(f"  Avg Confidence:       {stats['avg_confidence']:.1%}")
    print(f"  High Confidence (90%+): {stats['high_confidence_count']}")
    print("\n  Mapping Types:")
    for mtype, count in stats['mapping_types'].items():
        print(f"    {mtype}: {count}")
    print("\nOutputs:")
    print(f"  {config.out_lineage_dir / 'table_lineage.csv'}")
    print(f"  {config.out_lineage_dir / 'attribute_lineage.csv'}")
    print(f"  {config.out_lineage_dir / 'table_lineage.xlsx'}")
    print(f"  {config.out_lineage_dir / 'attribute_lineage.xlsx'}")
    print("=" * 70)


if __name__ == "__main__":
    main()
